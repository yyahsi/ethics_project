from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO, Dict, Iterable, Mapping, MutableMapping, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

COMPARISON_METRICS = ["Ethicality", "Visibility", "Accountability"]
TITLE_COLUMN = "News Title"
COMPARISON_SUBCOLUMNS = ["Israel", "Russia"]
# Output layout: one Israel title column plus paired Israel/Russia metric columns.
OUTPUT_COLUMN_COUNT = 1 + len(COMPARISON_METRICS) * len(COMPARISON_SUBCOLUMNS)
DEFAULT_MODELS = ["ChatGPT", "Claude", "Gemini"]
DEFAULT_MAX_ROWS = 200
INCIDENT_EVALUATORS = ["My Ground Truth", "ChatGPT", "Claude", "Gemini"]

TITLE_CANDIDATES = [
    "news title",
    "headline",
    "news item",
    "title",
]


def clean_cell(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip().replace("\ufeff", "")
    text = text.replace("\\,", ",")
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        text = text[1:-1]
    elif text.startswith('"'):
        text = text[1:]
    elif text.endswith('"'):
        text = text[:-1]
    return text.replace('""', '"').strip()


def clean_metric(value: object):
    text = clean_cell(value)
    if text == "" or text.lower() in {"none", "nan", "null", "<na>"}:
        return pd.NA
    match = re.search(r"-?\d+", text)
    if not match:
        return pd.NA
    number = int(match.group())
    if 0 <= number <= 10:
        return number
    return pd.NA


def _read_text_from_path_or_buffer(path_or_buffer) -> str:
    if hasattr(path_or_buffer, "seek"):
        path_or_buffer.seek(0)
        raw = path_or_buffer.read()
        if isinstance(raw, bytes):
            return raw.decode("utf-8-sig", errors="replace")
        return str(raw)
    return Path(path_or_buffer).read_text(encoding="utf-8-sig", errors="replace")


def parse_loose_metric_csv_text(text: str) -> pd.DataFrame:
    """Parse loose CSV/TXT files where headlines may contain unquoted commas.

    The parser assumes the last three comma-separated fields are:
    Ethicality, Visibility, Accountability. Everything before them is treated as
    an optional ID plus the title/headline.
    """
    rows = []
    lines = [line.rstrip("\r\n") for line in text.splitlines() if line.strip()]
    for line in lines[1:]:
        parts = line.rsplit(",", 3)
        if len(parts) != 4:
            continue
        left, ethicality, visibility, accountability = parts
        left = clean_cell(left)
        # Drop a leading ID if the row starts as "123,Title".
        if "," in left:
            possible_id, title = left.split(",", 1)
            if re.fullmatch(r"\s*#?\d+\s*", clean_cell(possible_id)):
                left = title
        title = clean_cell(left)
        if not title:
            continue
        rows.append(
            {
                "News Title": title,
                "Ethicality": clean_metric(ethicality),
                "Visibility": clean_metric(visibility),
                "Accountability": clean_metric(accountability),
            }
        )
    return pd.DataFrame(rows, columns=["News Title", *COMPARISON_METRICS])


def _find_title_column(columns: Iterable[object]) -> str:
    clean_columns = [clean_cell(c) for c in columns]
    lower_map = {c.lower(): c for c in clean_columns}
    for candidate in TITLE_CANDIDATES:
        if candidate in lower_map:
            return lower_map[candidate]
    for c in clean_columns:
        lower = c.lower()
        if "headline" in lower or ("news" in lower and "title" in lower) or "item" in lower:
            return c
    raise ValueError("Missing title/headline column.")


def _find_metric_column(columns: Iterable[object], metric: str, model_hint: str | None = None) -> str:
    clean_columns = [clean_cell(c) for c in columns]
    lower_columns = [(c.lower(), c) for c in clean_columns]
    metric_lower = metric.lower()
    model_lower = clean_cell(model_hint).lower() if model_hint else ""

    # 1) Exact metric name, e.g. Ethicality.
    for lower, col in lower_columns:
        if lower == metric_lower:
            return col

    # 2) Model-prefixed metric, e.g. ChatGPT Ethicality.
    if model_lower:
        for lower, col in lower_columns:
            if model_lower in lower and metric_lower in lower:
                return col

    # 3) Common normalized variants, e.g. ethicality_0_10.
    for lower, col in lower_columns:
        if metric_lower in lower:
            return col

    raise ValueError(f"Missing {metric} column.")


def normalise_comparison_table(df: pd.DataFrame, model_hint: str | None = None) -> pd.DataFrame:
    """Return columns: News Title, Ethicality, Visibility, Accountability."""
    if df.empty:
        return pd.DataFrame(columns=["News Title", *COMPARISON_METRICS])

    df = df.copy()
    df.columns = [clean_cell(c) for c in df.columns]

    title_col = _find_title_column(df.columns)
    metric_cols = {metric: _find_metric_column(df.columns, metric, model_hint) for metric in COMPARISON_METRICS}

    out = pd.DataFrame()
    out["News Title"] = df[title_col].map(clean_cell)
    for metric, source_col in metric_cols.items():
        out[metric] = df[source_col].map(clean_metric).astype("Int64")

    out = out[out["News Title"].ne("")].reset_index(drop=True)
    return out


def read_comparison_file(path_or_buffer, model_hint: str | None = None) -> pd.DataFrame:
    """Read CSV, TXT, XLS, or XLSX files and normalize comparison columns."""
    name = getattr(path_or_buffer, "name", str(path_or_buffer)).lower()
    if hasattr(path_or_buffer, "seek"):
        path_or_buffer.seek(0)

    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(path_or_buffer)
        return normalise_comparison_table(df, model_hint=model_hint)

    # First try normal CSV parsing, which preserves extra fields such as rationale.
    try:
        if hasattr(path_or_buffer, "seek"):
            path_or_buffer.seek(0)
        df = pd.read_csv(path_or_buffer, encoding="utf-8-sig")
        return normalise_comparison_table(df, model_hint=model_hint)
    except Exception:
        text = _read_text_from_path_or_buffer(path_or_buffer)
        df = parse_loose_metric_csv_text(text)
        return normalise_comparison_table(df, model_hint=model_hint)


def make_side_by_side_dataframe(news1: pd.DataFrame, news2: pd.DataFrame, max_rows: int = DEFAULT_MAX_ROWS) -> pd.DataFrame:
    """Create a pandas table with the final side-by-side output layout.

    The title column comes only from the Israel file. The three metric groups keep
    paired Israel/Russia subcolumns.
    """
    news1 = normalise_comparison_table(news1).reset_index(drop=True)
    news2 = normalise_comparison_table(news2).reset_index(drop=True)
    records = []
    for idx in range(max_rows):
        row = []
        if idx < len(news1):
            value = news1.at[idx, TITLE_COLUMN]
            row.append("" if pd.isna(value) else value)
        else:
            row.append("")

        for metric in COMPARISON_METRICS:
            for source_df in (news1, news2):
                if idx < len(source_df):
                    value = source_df.at[idx, metric]
                    row.append("" if pd.isna(value) else value)
                else:
                    row.append("")
        records.append(row)

    columns = pd.MultiIndex.from_tuples(
        [(TITLE_COLUMN, "Israel")]
        + [(metric, source) for metric in COMPARISON_METRICS for source in COMPARISON_SUBCOLUMNS]
    )
    return pd.DataFrame(records, columns=columns)


def _safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", " ", clean_cell(name)).strip() or "Comparison"
    return safe[:31]


def _cell_value(value: object):
    if value is None or pd.isna(value):
        return ""
    return value


def make_average_scores_dataframe(
    pairs: Mapping[str, Tuple[pd.DataFrame, pd.DataFrame]],
    max_rows: int = DEFAULT_MAX_ROWS,
) -> pd.DataFrame:
    """Create one average-score table for all AI models.

    Output layout mirrors the comparison sheets: each metric has Israel and
    Russia subcolumns, but each row is a model average instead of an individual
    news item. Blank or invalid scores are ignored.
    """
    records = []
    ordered_models = [model for model in DEFAULT_MODELS if model in pairs]
    ordered_models += [model for model in pairs if model not in ordered_models]

    for model_name in ordered_models:
        news1, news2 = pairs[model_name]
        news1 = normalise_comparison_table(news1).head(max_rows).reset_index(drop=True)
        news2 = normalise_comparison_table(news2).head(max_rows).reset_index(drop=True)
        row = [model_name]
        for metric in COMPARISON_METRICS:
            for source_df in (news1, news2):
                average = source_df[metric].astype("Float64").mean(skipna=True)
                row.append("" if pd.isna(average) else round(float(average), 2))
        records.append(row)

    columns = pd.MultiIndex.from_tuples(
        [("AI Model", "")]
        + [(metric, source) for metric in COMPARISON_METRICS for source in COMPARISON_SUBCOLUMNS]
    )
    return pd.DataFrame(records, columns=columns)


def flatten_average_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a CSV-friendly copy of the average-score table."""
    out = df.copy()
    if isinstance(out.columns, pd.MultiIndex):
        out.columns = [" ".join(str(part) for part in col if str(part).strip()).strip() for col in out.columns]
    return out


def make_incident_average_scores_dataframe(ratings: pd.DataFrame) -> pd.DataFrame:
    """Create average incident-rating scores for ground truth and AI columns.

    The input is expected to contain the incident ratings table columns:
    My Ground Truth, ChatGPT, Claude, and Gemini. Blank or invalid scores are
    ignored when calculating averages.
    """
    ratings = ratings.copy()
    records = []
    for evaluator in INCIDENT_EVALUATORS:
        if evaluator not in ratings.columns:
            average = pd.NA
            count = 0
        else:
            numeric_scores = pd.to_numeric(ratings[evaluator], errors="coerce")
            average = numeric_scores.mean(skipna=True)
            count = int(numeric_scores.notna().sum())
        records.append(
            [
                evaluator,
                "" if pd.isna(average) else round(float(average), 2),
                count,
            ]
        )

    columns = pd.MultiIndex.from_tuples(
        [("Evaluator", ""), ("Average Score", "Incident Ratings"), ("Marked Rows", "Count")]
    )
    return pd.DataFrame(records, columns=columns)


def _write_incident_average_table(
    ws,
    table: pd.DataFrame,
    start_row: int,
    dark_fill: PatternFill,
    sub_fill: PatternFill,
    white_font: Font,
    bold_font: Font,
    border: Border,
    center: Alignment,
    left_wrap: Alignment,
) -> int:
    """Write the incident averages as a second styled table. Returns next row."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    title_cell = ws.cell(row=start_row, column=1, value="Incident Ratings Average Scores")
    title_cell.fill = dark_fill
    title_cell.font = white_font
    title_cell.alignment = center
    title_cell.border = border
    for col_idx in range(2, 4):
        ws.cell(row=start_row, column=col_idx).fill = dark_fill
        ws.cell(row=start_row, column=col_idx).border = border

    header_row = start_row + 1
    subheader_row = start_row + 2
    headers = [("Evaluator", "Evaluator"), ("Average Score", "Incident Ratings"), ("Marked Rows", "Count")]
    for col_idx, (header, subheader) in enumerate(headers, start=1):
        top = ws.cell(row=header_row, column=col_idx, value=header)
        top.fill = dark_fill
        top.font = white_font
        top.alignment = center
        top.border = border
        sub = ws.cell(row=subheader_row, column=col_idx, value=subheader)
        sub.fill = sub_fill
        sub.font = bold_font
        sub.alignment = center
        sub.border = border

    for row_idx, record in enumerate(flatten_average_columns(table).itertuples(index=False, name=None), start=start_row + 3):
        for col_idx, value in enumerate(record, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_wrap if col_idx == 1 else center
            if col_idx == 2:
                cell.number_format = "0.00"
            elif col_idx == 3:
                cell.number_format = "0"
    return start_row + 3 + len(table)


def create_side_by_side_workbook(
    pairs: Mapping[str, Tuple[pd.DataFrame, pd.DataFrame]],
    source_info: Mapping[str, Mapping[str, object]] | None = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    incident_ratings: pd.DataFrame | None = None,
) -> bytes:
    """Create an XLSX workbook with one side-by-side sheet per pair."""
    source_info = source_info or {}
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    summary_fill = PatternFill("solid", fgColor="E2F0D9")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    summary_headers = ["Comparison", "Israel file", "Israel rows", "Russia file", "Russia rows", "Output rows"]
    summary.append(summary_headers)
    for cell in summary[1]:
        cell.fill = summary_fill
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    average_table = make_average_scores_dataframe(pairs, max_rows=max_rows)
    secret_ws = wb.create_sheet("Final Comparison")
    secret_ws.freeze_panes = "A3"
    secret_ws.sheet_view.showGridLines = False

    secret_ws.cell(row=1, column=1, value="AI Model")
    secret_ws.cell(row=2, column=1, value="AI Model")
    secret_ws.cell(row=1, column=1).fill = dark_fill
    secret_ws.cell(row=1, column=1).font = white_font
    secret_ws.cell(row=1, column=1).alignment = center
    secret_ws.cell(row=1, column=1).border = border
    secret_ws.cell(row=2, column=1).fill = sub_fill
    secret_ws.cell(row=2, column=1).font = bold_font
    secret_ws.cell(row=2, column=1).alignment = center
    secret_ws.cell(row=2, column=1).border = border

    col = 2
    for metric in COMPARISON_METRICS:
        secret_ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = secret_ws.cell(row=1, column=col, value=metric)
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border
        secret_ws.cell(row=1, column=col + 1).fill = dark_fill
        secret_ws.cell(row=1, column=col + 1).border = border
        for offset, subheader in enumerate(COMPARISON_SUBCOLUMNS):
            sub_cell = secret_ws.cell(row=2, column=col + offset, value=subheader)
            sub_cell.fill = sub_fill
            sub_cell.font = bold_font
            sub_cell.alignment = center
            sub_cell.border = border
        col += 2

    for row_idx, record in enumerate(flatten_average_columns(average_table).itertuples(index=False, name=None), start=3):
        for col_idx, value in enumerate(record, start=1):
            cell = secret_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_wrap if col_idx == 1 else center
            if col_idx > 1:
                cell.number_format = "0.00"

    secret_ws.column_dimensions["A"].width = 18
    for col_idx in range(2, OUTPUT_COLUMN_COUNT + 1):
        secret_ws.column_dimensions[get_column_letter(col_idx)].width = 15
    secret_ws.row_dimensions[1].height = 24
    secret_ws.row_dimensions[2].height = 22

    if incident_ratings is not None:
        incident_table = make_incident_average_scores_dataframe(incident_ratings)
        second_table_start = 3 + len(average_table) + 2
        _write_incident_average_table(
            secret_ws,
            incident_table,
            second_table_start,
            dark_fill,
            sub_fill,
            white_font,
            bold_font,
            border,
            center,
            left_wrap,
        )
        secret_ws.column_dimensions["A"].width = max(secret_ws.column_dimensions["A"].width or 0, 20)
        secret_ws.column_dimensions["B"].width = max(secret_ws.column_dimensions["B"].width or 0, 18)
        secret_ws.column_dimensions["C"].width = max(secret_ws.column_dimensions["C"].width or 0, 14)

    for model_name, (news1, news2) in pairs.items():
        news1 = normalise_comparison_table(news1).reset_index(drop=True)
        news2 = normalise_comparison_table(news2).reset_index(drop=True)
        info = source_info.get(model_name, {})
        summary.append(
            [
                f"{model_name} Comparison",
                info.get("news1_file", ""),
                len(news1),
                info.get("news2_file", ""),
                len(news2),
                max_rows,
            ]
        )

        ws = wb.create_sheet(_safe_sheet_name(f"{model_name} Comparison"))
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

        # Top grouped headers. Column A is only the Israel title.
        title_header = ws.cell(row=1, column=1, value=TITLE_COLUMN)
        title_header.fill = dark_fill
        title_header.font = white_font
        title_header.alignment = center
        title_header.border = border

        col = 2
        for metric in COMPARISON_METRICS:
            start = col
            end = col + 1
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=metric)
            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border
            ws.cell(row=1, column=end).fill = dark_fill
            ws.cell(row=1, column=end).border = border
            col += 2

        # Inner headers.
        inner_headers = ["Israel"] + COMPARISON_SUBCOLUMNS * len(COMPARISON_METRICS)
        for idx, value in enumerate(inner_headers, start=1):
            cell = ws.cell(row=2, column=idx, value=value)
            cell.fill = sub_fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

        # Body.
        for row_idx in range(max_rows):
            values = []
            if row_idx < len(news1):
                values.append(_cell_value(news1.at[row_idx, TITLE_COLUMN]))
            else:
                values.append("")

            for metric in COMPARISON_METRICS:
                for source_df in (news1, news2):
                    if row_idx < len(source_df):
                        values.append(_cell_value(source_df.at[row_idx, metric]))
                    else:
                        values.append("")
            excel_row = row_idx + 3
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = left_wrap if col_idx == 1 else center

        ws.auto_filter.ref = f"A2:G{max_rows + 2}"
        ws.column_dimensions["A"].width = 58
        for col_idx in range(2, OUTPUT_COLUMN_COUNT + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        for row_idx in range(3, max_rows + 3):
            ws.row_dimensions[row_idx].height = 38

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = left_wrap if cell.column in (1, 2, 4) else center
    for idx, width in enumerate([24, 34, 12, 34, 12, 12], start=1):
        summary.column_dimensions[get_column_letter(idx)].width = width
    summary.freeze_panes = "A2"
    summary.sheet_view.showGridLines = False

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _first_matching_file(directory: Path, patterns: Iterable[str]) -> Path | None:
    for pattern in patterns:
        matches = sorted(p for p in directory.glob(pattern) if p.is_file())
        if matches:
            return matches[0]
    return None


def build_pairs_from_directory(input_dir: str | Path) -> tuple[dict[str, tuple[pd.DataFrame, pd.DataFrame]], dict[str, dict[str, object]]]:
    """Load the default ChatGPT/Claude/Gemini Israel-Russia files from a folder."""
    input_dir = Path(input_dir)
    patterns = {
        "ChatGPT": {
            "news1": ["chatgpt*news1*", "chatgpt*1*"],
            "news2": ["chatgpt*news2*", "chatgpt*2*"],
        },
        "Claude": {
            "news1": ["claude*news1*", "claude*1*"],
            "news2": ["claude*news2*", "claude*2*"],
        },
        "Gemini": {
            "news1": ["gemini*news1*", "gemini*1*"],
            "news2": ["gemini*news2*", "gemini*2*"],
        },
    }
    pairs: dict[str, tuple[pd.DataFrame, pd.DataFrame]] = {}
    info: dict[str, dict[str, object]] = {}
    missing = []
    for model, model_patterns in patterns.items():
        p1 = _first_matching_file(input_dir, model_patterns["news1"])
        p2 = _first_matching_file(input_dir, model_patterns["news2"])
        if p1 is None or p2 is None:
            missing.append(model)
            continue
        df1 = read_comparison_file(p1, model_hint=model)
        df2 = read_comparison_file(p2, model_hint=model)
        pairs[model] = (df1, df2)
        info[model] = {
            "news1_file": p1.name,
            "news2_file": p2.name,
        }
    if missing:
        raise FileNotFoundError(f"Missing Israel/Russia input files for: {', '.join(missing)}")
    return pairs, info
